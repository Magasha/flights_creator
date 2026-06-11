"""
Модуль для чтения Excel файла с данными по заказам и внешним исполнителям.
"""

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

import openpyxl

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------
# Ожидаемые названия колонок в Excel (регистронезависимо)
# ----------------------------------------------------------------
COLUMN_MAP = {
    # Данные заказа
    "order_id":         ["id заказа", "номер заказа", "order_id", "orderid", "id"],
    "flight_name":      ["название рейса", "рейс", "flight_name", "flightname"],

    # Внешний исполнитель
    "company_name":     ["наименование компании", "компания", "company_name", "company"],
    "inn":              ["инн", "inn", "тин", "tin"],
    "org_type":         ["тип компании", "org_type", "тип организации", "orgtype", "правовая форма"],
    "country":          ["страна", "country", "код страны", "country_code"],
    "currency":         ["валюта", "currency"],
    "vat":              ["налоговая ставка", "ндс %", "ндс", "vat", "налог"],
    "price":            ["стоимость без ндс", "цена без ндс", "price", "стоимость"],
    "price_with_vat":   ["стоимость с ндс", "цена с ндс", "price_with_vat", "итого"],

    # Опциональные поля
    "contact_name":     ["контактное лицо", "фио", "contact_name", "контакт"],
    "contact_phone":    ["телефон", "phone", "contact_phone"],
    "contact_email":    ["email", "почта", "contact_email", "e-mail"],
    "comment":          ["комментарий", "comment", "примечание"],
    "is_vat_payer":     ["плательщик ндс", "is_vat_payer", "ндс плательщик"],
}


@dataclass
class OrderRow:
    """Одна строка из Excel — один заказ для обработки."""
    order_id: str
    flight_name: str

    # Внешний исполнитель (обязательные)
    company_name: str
    inn: str
    org_type: str
    country: str
    currency: str
    vat: float
    price: float
    price_with_vat: float

    # Опциональные
    contact_name: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    comment: Optional[str] = None
    is_vat_payer: bool = True

    # Системные
    row_number: int = 0
    errors: List[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0


def _normalize(value) -> str:
    """Приводит значение ячейки к строке, обрезая пробелы."""
    if value is None:
        return ""
    return str(value).strip()


def _normalize_lower(value) -> str:
    return _normalize(value).lower()


def _find_column_index(headers: list, field_key: str) -> Optional[int]:
    """
    Ищет индекс колонки по списку допустимых названий.
    Возвращает None если колонка не найдена.
    """
    aliases = COLUMN_MAP.get(field_key, [])
    for idx, header in enumerate(headers):
        if _normalize_lower(header) in aliases:
            return idx
    return None


def _parse_bool(value) -> bool:
    """Парсит булево значение из ячейки."""
    s = _normalize_lower(value)
    if s in ("да", "yes", "true", "1", "+"):
        return True
    if s in ("нет", "no", "false", "0", "-"):
        return False
    return True  # по умолчанию — плательщик НДС


def _parse_float(value, field_name: str, row_errors: list) -> float:
    """Парсит число из ячейки. При ошибке добавляет сообщение в row_errors."""
    s = _normalize(value).replace(",", ".").replace(" ", "").replace("\xa0", "")
    if not s:
        row_errors.append(f"Поле '{field_name}' пустое")
        return 0.0
    try:
        return float(s)
    except ValueError:
        row_errors.append(f"Поле '{field_name}': некорректное число '{value}'")
        return 0.0


def read_excel(file_path: str, sheet_name: str = "Рейсы") -> List[OrderRow]:
    """
    Читает Excel файл и возвращает список OrderRow.

    Args:
        file_path:   Путь к .xlsx файлу
        sheet_name:  Название листа (по умолчанию 'Рейсы')

    Returns:
        Список OrderRow (включая строки с ошибками — проверяйте .is_valid)

    Raises:
        FileNotFoundError: Файл не найден
        ValueError: Лист не найден или не обнаружена строка заголовков
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    logger.info("Открываю файл: %s, лист: %s", file_path, sheet_name)

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Лист '{sheet_name}' не найден. Доступные листы: {available}"
        )

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Лист пустой — нет строк с данными")

    # Ищем строку заголовков (первую непустую)
    header_row_idx = None
    headers = []
    for i, row in enumerate(rows):
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 3:
            header_row_idx = i
            headers = [_normalize(c) for c in row]
            break

    if header_row_idx is None:
        raise ValueError("Не удалось найти строку с заголовками (минимум 3 заполненных ячейки)")

    logger.debug("Заголовки найдены в строке %d: %s", header_row_idx + 1, headers)

    # Строим карту: поле → индекс колонки
    col_idx = {}
    for field_key in COLUMN_MAP:
        col_idx[field_key] = _find_column_index(headers, field_key)

    # Проверяем обязательные поля
    required_fields = [
        "order_id", "flight_name", "company_name", "inn",
        "org_type", "country", "currency", "vat", "price", "price_with_vat"
    ]
    missing_required = [f for f in required_fields if col_idx.get(f) is None]
    if missing_required:
        raise ValueError(
            f"В заголовках не найдены обязательные колонки: {missing_required}\n"
            f"Фактические заголовки: {headers}\n"
            f"Допустимые варианты названий — см. README.md"
        )

    result = []
    data_rows = rows[header_row_idx + 1:]

    for local_idx, row in enumerate(data_rows):
        row_number = header_row_idx + local_idx + 2  # номер строки в Excel (1-based)

        # Пропускаем полностью пустые строки
        if all(c is None or _normalize(c) == "" for c in row):
            continue

        def get(field_key: str) -> str:
            idx = col_idx.get(field_key)
            if idx is None or idx >= len(row):
                return ""
            return _normalize(row[idx])

        errors = []

        order_id = get("order_id")
        flight_name = get("flight_name")
        company_name = get("company_name")
        inn = get("inn")
        org_type = get("org_type")
        country = get("country")
        currency = get("currency")
        vat_raw = get("vat")
        price_raw = get("price")
        price_with_vat_raw = get("price_with_vat")

        # Валидация обязательных строковых полей
        for fname, fval in [
            ("ID заказа", order_id),
            ("Название рейса", flight_name),
            ("Наименование компании", company_name),
            ("ИНН", inn),
            ("Тип компании", org_type),
            ("Страна", country),
            ("Валюта", currency),
        ]:
            if not fval:
                errors.append(f"Пустое обязательное поле: '{fname}'")

        vat = _parse_float(vat_raw, "Налоговая ставка НДС", errors)
        price = _parse_float(price_raw, "Стоимость без НДС", errors)
        price_with_vat = _parse_float(price_with_vat_raw, "Стоимость с НДС", errors)

        # Опциональные поля
        contact_name = get("contact_name") or None
        contact_phone = get("contact_phone") or None
        contact_email = get("contact_email") or None
        comment = get("comment") or None
        is_vat_payer_raw = get("is_vat_payer")
        is_vat_payer = _parse_bool(is_vat_payer_raw) if is_vat_payer_raw else True

        order_row = OrderRow(
            order_id=order_id,
            flight_name=flight_name,
            company_name=company_name,
            inn=inn,
            org_type=org_type,
            country=country,
            currency=currency,
            vat=vat,
            price=price,
            price_with_vat=price_with_vat,
            contact_name=contact_name,
            contact_phone=contact_phone,
            contact_email=contact_email,
            comment=comment,
            is_vat_payer=is_vat_payer,
            row_number=row_number,
            errors=errors,
        )
        result.append(order_row)

    logger.info("Прочитано строк: %d (включая %d с ошибками)",
                len(result), sum(1 for r in result if not r.is_valid))
    wb.close()
    return result
